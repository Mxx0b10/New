import os
import io
import json
import base64
import tempfile
import re
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = tempfile.mkdtemp()
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))


def style_worksheet(ws, headers_row=1):
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_idx == headers_row:
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


def pdf_to_excel(file_bytes):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tables_found = 0

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            if tables:
                for tbl_idx, table in enumerate(tables):
                    if not table:
                        continue
                    sheet_name = f"Page{page_num}_T{tbl_idx+1}"[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    for row in table:
                        ws.append([cell if cell is not None else "" for cell in row])
                    style_worksheet(ws)
                    tables_found += 1
            else:
                # Extract raw text as fallback
                text = page.extract_text()
                if text and text.strip():
                    sheet_name = f"Page{page_num}_Text"[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    for line in text.split("\n"):
                        ws.append([line.strip()])
                    style_worksheet(ws)
                    tables_found += 1

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    return wb, tables_found


def image_to_excel_via_claude(file_bytes, media_type):
    b64 = base64.standard_b64encode(file_bytes).decode("utf-8")

    prompt = """Analyze this image carefully and extract ALL tabular data, numbers, and structured information.

Return ONLY a valid JSON object in this exact format (no markdown, no explanation):
{
  "sheets": [
    {
      "name": "Sheet name (descriptive)",
      "headers": ["col1", "col2", ...],
      "rows": [
        ["value1", "value2", ...],
        ...
      ]
    }
  ],
  "summary": "Brief description of what was extracted"
}

Rules:
- Extract every table, list, or structured data visible
- If no clear table exists, organize text data into logical columns
- Preserve numbers exactly as shown
- Use multiple sheets if multiple distinct tables exist
- For forms/invoices, create sheets for each section
- Never return empty sheets"""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    raw = message.content[0].text.strip()
    # Strip markdown code fences if present
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    data = json.loads(raw)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_data in data.get("sheets", []):
        name = sheet_data.get("name", "Sheet")[:31]
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        ws = wb.create_sheet(title=name)
        if headers:
            ws.append(headers)
        for row in rows:
            ws.append([str(v) if v is not None else "" for v in row])
        style_worksheet(ws)

    if not wb.sheetnames:
        wb.create_sheet("Extracted")

    return wb, data.get("summary", "Data extracted successfully")


def pdf_with_claude_fallback(file_bytes):
    """Try pdfplumber first; if minimal data, enhance with Claude vision."""
    wb, tables_found = pdf_to_excel(file_bytes)

    if tables_found < 1:
        # Convert first page to image and use Claude
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            page = doc[0]
            pix = page.get_pixmap(dpi=150)
            img_bytes = pix.tobytes("png")
            wb, summary = image_to_excel_via_claude(img_bytes, "image/png")
            return wb, f"Claude Vision: {summary}"
        except ImportError:
            pass

    return wb, f"Extracted {tables_found} table(s) from PDF"


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/convert", methods=["POST"])
def convert():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    file_bytes = file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        return jsonify({"error": "File too large (max 20MB)"}), 400

    filename = file.filename.lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""

    try:
        if ext == "pdf":
            wb, summary = pdf_with_claude_fallback(file_bytes)
        elif ext in ("png", "jpg", "jpeg", "webp", "gif", "bmp", "tiff", "tif"):
            media_map = {
                "jpg": "image/jpeg", "jpeg": "image/jpeg",
                "png": "image/png", "webp": "image/webp",
                "gif": "image/gif",
            }
            media_type = media_map.get(ext, "image/png")
            wb, summary = image_to_excel_via_claude(file_bytes, media_type)
        else:
            return jsonify({"error": "Unsupported file type. Use PDF, PNG, JPG, or WEBP."}), 400

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        base_name = file.filename.rsplit(".", 1)[0]
        output_name = f"{base_name}_converted.xlsx"

        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=output_name,
        )

    except json.JSONDecodeError as e:
        return jsonify({"error": f"Claude returned invalid JSON: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000)
